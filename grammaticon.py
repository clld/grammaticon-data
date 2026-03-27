#!/usr/bin/env python3

import csv
import io
import json
import platform
import re
import shutil
import subprocess
import sys
import zipfile
from collections import defaultdict
from contextlib import ExitStack
from itertools import chain, repeat
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.parse import quote

# dependencies for xlsx-to-csv
try:
    XLSX_TO_CSV_DEPS = ['openpyxl']
    from openpyxl import load_workbook
    xlsx_to_csv_deps_okay = True
except ModuleNotFoundError:
    xlsx_to_csv_deps_okay = False

# dependencies for make-csvw
try:
    MAKE_CSVW_DEPS = ['csvw', 'simplepybtex']
    from csvw.metadata import Column, Table, TableGroup
    from simplepybtex.database import BibliographyData, parse_file
    make_csvw_deps_okay = True
except ModuleNotFoundError:
    make_csvw_deps_okay = False

USAGE = """usage: {progname} command [options]

supported commands

\txlsx-to-csv
\t\tconvert excel spread sheets in raw/ to csv files in raw/csv-export/
\tdownload-collections
\t\tdownload cldf versions of the collections into raw/download/
\tmake-csvw
\t\tcreate CSVW dataset in csvw/
\t-h, --help
\t\tprint this message"""

HERE = Path(__file__).parent
RAW_DIR = HERE / 'raw'
CSV_DIR = RAW_DIR / 'csv-export'
DOWNLOAD_DIR = RAW_DIR / 'download'
DEST_DIR = HERE / 'csvw'

# Conversion from Excel to CSV

def normalise_excel_cell(value):
    if value is None:
        return ''
    else:
        return str(value).strip()


def pad_list(ls, width):
    if len(ls) == width:
        return ls
    elif len(ls) < width:
        return list(chain(ls, repeat('', width - len(ls))))
    else:
        raise ValueError(f'too long: {ls}')


def xlsx_file_to_csv_file(excel_path, outdir):
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    worksheets = wb.worksheets
    assert len(worksheets) == 1, f'{excel_path}: not exactly 1 worksheet'
    sheet = worksheets[0]
    rows = [
        row_norm
        for row in sheet.iter_rows()
        if any(row_norm := [normalise_excel_cell(cell.value) for cell in row])]

    dest = outdir.joinpath(excel_path.name).with_suffix('.csv')
    table_width = max(len(row) for row in rows)
    if not outdir.is_dir():
        outdir.mkdir()
    with open(dest, 'w', encoding='utf-8') as f:
        wtr = csv.writer(f)
        wtr.writerows(pad_list(row, table_width) for row in rows)


def xlsx_to_csv():
    if not xlsx_to_csv_deps_okay:
        print('the make-csvw command requires following python packages:', file=sys.stderr)
        print('\n'.join(f'\t{dep}' for dep in XLSX_TO_CSV_DEPS), file=sys.stderr)
        sys.exit(72)
    for p in RAW_DIR.glob('*.xlsx'):
        xlsx_file_to_csv_file(p, CSV_DIR)


# Downloading the collections

def read_csv(f):
    rdr = csv.reader(f)
    header = next(rdr)
    for row in rdr:
        yield {k: v.strip() for k, v in zip(header, row) if v.strip()}


def get_zenodo_no(doi):
    if (m := re.fullmatch(r'10\.5281/zenodo\.(\d+)', doi)):
        return int(m.group(1))
    else:
        msg = 'doi looks funky: {}'.format(doi)
        raise AssertionError(msg)


def get_zip_path(record_no):
    return DOWNLOAD_DIR / f'{record_no}.zip'


def download_collections():
    with open(RAW_DIR / 'dois.csv', encoding='utf-8') as f:
        collections = list(read_csv(f))
    for coll in collections:
        coll['ID'] = get_zenodo_no(coll['DOI'])
        coll['Zip_Path'] = get_zip_path(coll['ID'])
    collections = {coll['ID']: coll for coll in collections}

    missing_records = [
        record_no
        for record_no, coll in collections.items()
        if not coll['Zip_Path'].exists()]
    if not missing_records:
        print('Nothing to do.', file=sys.stderr)
        return

    query = 'OR'.join(
        '(id:{})'.format(record_no) for record_no in missing_records)
    zenodo_url = f'https://zenodo.org/api/records?q={quote(query)}'
    assert zenodo_url.startswith('https://')
    req = Request(zenodo_url, headers={'Content-Type': 'application/json'})
    with urlopen(req) as resp:
        record_metadata = json.load(resp)
    # TODO: worry about pagination later
    assert len(record_metadata['hits']['hits']) == len(missing_records), "it's time to worry about pagination"

    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    for zenodo_record in record_metadata['hits']['hits']:
        record_no = zenodo_record['id']
        assert len(zenodo_record['files']) == 1, record_no
        out_path = collections[record_no]['Zip_Path']
        print(f'downloading {out_path}...', file=sys.stderr)
        zip_url = zenodo_record['files'][0]['links']['self']
        assert zip_url.startswith('https://')
        with urlopen(zip_url) as resp, open(out_path, 'wb') as f:
            while (chunk := resp.read(4096)):
                f.write(chunk)


# CSVW creation

PROP_STRUCTURE_DATASET = 'http://cldf.clld.org/v1.0/terms.rdf#StructureDataset'

PROP_PARAMETER_TABLE = 'http://cldf.clld.org/v1.0/terms.rdf#ParameterTable'
PROP_VALUE_TABLE = 'http://cldf.clld.org/v1.0/terms.rdf#ValueTable'

PROP_ID = 'http://cldf.clld.org/v1.0/terms.rdf#id'
PROP_NAME = 'http://cldf.clld.org/v1.0/terms.rdf#name'
PROP_DESCRIPTION = 'http://cldf.clld.org/v1.0/terms.rdf#description'
PROP_PARAMETER_ID = 'http://cldf.clld.org/v1.0/terms.rdf#parameterReference'
PROP_LANGUAGE_ID = 'http://cldf.clld.org/v1.0/terms.rdf#languageReference'

RAW_TO_CSWV_MAP = {
    'Concepts.csv': {
        'name': 'concepts.csv',
        'columns': {
            'id': {
                'name': 'ID',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#id'},
            'label': {
                'name': 'Name',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#name'},
            'grammacode': {
                'name': 'Grammacode',
                'datatype': 'string'},
            'definition': {
                'name': 'Description',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#definition'},
            'comments': {
                'name': 'Comment',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#comment'},
            'SIL counterpart': {
                'name': 'SIL_Counterpart',
                'datatype': 'string'},
            'SIL URL': {
                'name': 'SIL_URL',
                'datatype': 'string'},
            'Wikipedia counterpart': {
                'name': 'Wikipedia_Counterpart',
                'datatype': 'string'},
            'Wikipedia URL': {
                'name': 'Wikipedia_URL',
                'datatype': 'string'},
            'Croft counterpart': {
                'name': 'Croft_counterpart',
                'datatype': 'string'},
            'Croft definition': {
                'name': 'Croft_definition',
                'datatype': 'string'},
            'Croft URL': {
                'name': 'Croft_URL',
                'datatype': 'string'},
            'GOLD counterpart': {
                'name': 'GOLD_counterpart',
                'datatype': 'string'},
            'ISOCAT counterpart': {
                'name': 'ISOCAT_counterpart',
                'datatype': 'string'},
            'quotation': {
                'name': 'Quotation',
                'datatype': 'string'},
            'Bibsources': {
                'name': 'Source',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#source',
                'separator': ';'}}},

    'Feature_lists.csv': {
        'name': 'collections.csv',
        'properties': {
            'dc:conformsTo': 'http://cldf.clld.org/v1.0/terms.rdf#ContributionTable',
        },
        'columns': {
            'id': {
                'name': 'ID',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#id'},
            'name': {
                'name': 'Name',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#name'},
            'URL': {
                'name': 'URL',
                'datatype': 'string'},
            'description': {
                'name': 'Description',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#description'},
            'year': {
                'name': 'Year',
                'datatype': 'string'},
            'Collection contributors':  {
                'name': 'Contributors',
                'datatype': 'string'}}},

    'Features.csv': {
        'name': 'features.csv',
        'columns': {
            'feature_ID': {
                'name': 'ID',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#id'},
            'feature name': {
                'name': 'Name',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#name'},
            'feature description': {
                'name': 'Description',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#definition'},
            'meta_feature_id': {
                'name': 'Metafeature_ID',
                'datatype': 'string'},
            'collection_id': {
                'name': 'Collection_ID',
                'datatype': 'string'},
            'feature URL': {
                'name': 'Feature_URL',
                'datatype': 'string'},
            'ID_in_collection': {
                'name': 'ID_in_Collection',
                'datatype': 'string'},
            'number of languages': {
                'name': 'Language_Count',
                'datatype': 'integer'},
            'comments': {
                'name': 'Comment',
                'datatype': 'string',
                'propertyUrl': 'http://cldf.clld.org/v1.0/terms.rdf#comment'}},
        'foreign-keys': {
            'Collection_ID': 'collections.csv'}},

    'Concepts_features.csv': {
        'name': 'concepts-features.csv',
        'columns': {
            'concept_id': {
                'name': 'Concept_ID',
                'datatype': 'string'},
            'feature_id': {
                'name': 'Feature_ID',
                'datatype': 'string'},
            'Comment': {
                'name': 'Comment',
                'datatype': 'string'}},
        'foreign-keys': {
            'Concept_ID': 'concepts.csv',
            'Feature_ID': 'features.csv'}}}


CONCEPT_ID_COL = 'concept_id'
CHILD_COL = 'concept_child_id'
PARENT_COL = 'concept_parent_id'

BIBKEY_FIXES = {
    'blomfield_language_1933': 'bloomfield_language_1933',
    'croft_morphosyntax_ 2022': 'croft_morphosyntax_2022',
}


def get_collection_parameters_from_zip(path):
    parameters = {}
    with ExitStack() as stack:
        zf = stack.enter_context(zipfile.ZipFile(path))
        json_files = [
            info for info in zf.infolist() if info.filename.endswith('.json')]
        json_metadata = [
            (info, md)
            for info in json_files
            if (md := json.load(stack.enter_context(zf.open(info))))
            and isinstance(md, dict)
            and md.get('dc:conformsTo') == PROP_STRUCTURE_DATASET]
        for info, md in json_metadata:
            parameter_table_name = None
            parameter_id_col = None
            parameter_name_col = None
            parameter_desc_col = None
            value_table_name = None
            value_parameter_col = None
            value_language_col = None
            for table in md['tables']:
                if table.get('dc:conformsTo') == PROP_PARAMETER_TABLE:
                    parameter_table_name = table.get('url')
                    for colspec in table['tableSchema']['columns']:
                        if colspec.get('propertyUrl') == PROP_ID:
                            parameter_id_col = colspec['name']
                        elif colspec.get('propertyUrl') == PROP_NAME:
                            parameter_name_col = colspec['name']
                        elif colspec.get('propertyUrl') == PROP_DESCRIPTION:
                            parameter_desc_col = colspec['name']
                elif table.get('dc:conformsTo') == PROP_VALUE_TABLE:
                    value_table_name = table.get('url')
                    for colspec in table['tableSchema']['columns']:
                        if colspec.get('propertyUrl') == PROP_PARAMETER_ID:
                            value_parameter_col = colspec['name']
                        elif colspec.get('propertyUrl') == PROP_LANGUAGE_ID:
                            value_language_col = colspec['name']
            if parameter_table_name is None:
                continue

            cldf_path = Path(info.filename).parent

            languages_per_parameter_id = defaultdict(set)
            if value_table_name:
                vf = stack.enter_context(zf.open(str(cldf_path / parameter_table_name)))
                vf_unicode = io.TextIOWrapper(vf, encoding='utf-8')
                for row in read_csv(vf_unicode):
                    parameter_id = row.get(value_parameter_col)
                    language_id = row.get(value_language_col)
                    if parameter_id and language_id:
                        languages_per_parameter_id[parameter_id].add(language_id)
            language_counts = {id_: len(lgs) for id_, lgs in languages_per_parameter_id.items()}

            pf = stack.enter_context(zf.open(str(cldf_path / parameter_table_name)))
            pf_unicode = io.TextIOWrapper(pf, encoding='utf-8')
            parameters.update(
                (parameter_id,
                 {'ID': parameter_id,
                  'Name': row.get(parameter_name_col) or '',
                  'Description': row.get(parameter_desc_col) or '',
                  'Language_Count': language_counts.get(parameter_id) or 0})
                for row in read_csv(pf_unicode)
                if (parameter_id := row.get(parameter_id_col)))
    return parameters


def simplified_concept_hierarchy(original_hierarchy, concept_ids):
    # The table looks like rows only have *either* a child_id *or* a parent id.
    # Check this assumption:
    conflicting = [
        row
        for row in original_hierarchy
        # source for the hacky xor: https://stackoverflow.com/a/433161
        if bool(row.get(CHILD_COL)) == bool(row.get(PARENT_COL))]
    assert not conflicting, f'concepthierarchy: concepts with parent *and* child: {conflicting}'

    # The table also looks reflexive. Every concept--child pair seems to have
    # a redundant concept--parent pair.
    # Check this assumption:
    children = {
        (row[CHILD_COL], row[CONCEPT_ID_COL])
        for row in original_hierarchy
        if row.get(CHILD_COL) in concept_ids
        and row.get(CONCEPT_ID_COL) in concept_ids}
    parents = {
        (row[CONCEPT_ID_COL], row[PARENT_COL])
        for row in original_hierarchy
        if row.get(PARENT_COL) in concept_ids
        and row.get(CONCEPT_ID_COL) in concept_ids}
    assert children == parents, 'I expect all pairs to be reflexive'

    def valid_hierarchy_path(child_id, parent_id):
        if child_id not in concept_ids:
            msg = (
                'concept-hierarchy.csv:'
                f" unknown child id '{child_id}' for parent {parent_id}")
            print(msg, file=sys.stderr)
            return False
        elif parent_id not in concept_ids:
            msg = (
                'concept-hierarchy.csv:'
                f" unknown parent id '{parent_id}' for child {child_id}")
            print(msg, file=sys.stderr)
            return False
        else:
            return True

    assocs = sorted(parents, key=lambda row: tuple(map(int, row)))
    return [
        {'Child_ID': child_id, 'Parent_ID': parent_id}
        for child_id, parent_id in assocs
        if valid_hierarchy_path(child_id, parent_id)]


def is_concept_valid(row):
    if 'Name' not in row:
        msg = (
            'concepts.csv:'
            ' missing name for concept {}'.format(row['ID']))
        print(msg, file=sys.stderr)
        return False
    else:
        return True


def only_valid_concepts(concepts):
    return list(filter(is_concept_valid, concepts))


def is_feature_valid(row, collection_ids):
    if 'Collection_ID' not in row:
        msg = 'features.csv: missing feature list id for feature {}'.format(
            row['ID'])
        print(msg, file=sys.stderr)
        return False
    elif (flid := row['Collection_ID']) not in collection_ids:
        msg = (
            'Features.csv:'
            ' invalid feature list id for feature {}: {}'.format(
                row['ID'], flid))
        print(msg, file=sys.stderr)
        return False
    else:
        return True


def only_valid_features(features, collection_ids):
    return [
        row
        for row in features
        if is_feature_valid(row, collection_ids)]


def is_concept_feature_valid(row, concept_ids, feature_ids):
    if 'Feature_ID' not in row:
        msg = (
            'concepts-features.csv:'
            ' missing feature id for concept {}'.format(
                row['Concept_ID']))
        print(msg, file=sys.stderr)
        return False
    elif (mfid := row['Feature_ID']) not in feature_ids:
        msg = (
            'concepts-features.csv:'
            ' invalid feature id for concept {}: {}'.format(
                row['Concept_ID'], mfid))
        print(msg, file=sys.stderr)
        return False
    elif (cid := row['Concept_ID']) not in concept_ids:
        msg = (
            'concepts-features.csv:'
            ' invalid concept id for feature {}: {}'.format(
                row['Feature_ID'], cid))
        print(msg, file=sys.stderr)
        return False
    else:
        return True


def only_valid_concept_features(concept_features, concept_ids, feature_ids):
    return [
        row
        for row in concept_features
        if is_concept_feature_valid(row, concept_ids, feature_ids)]


def make_csvw():
    if not make_csvw_deps_okay:
        print('the make-csvw command requires following python packages:', file=sys.stderr)
        print('\n'.join(f'\t{dep}' for dep in MAKE_CSVW_DEPS), file=sys.stderr)
        sys.exit(72)

    raw_tables = [
        CSV_DIR / 'Concepts.csv',
        CSV_DIR / 'Feature_lists.csv',
        CSV_DIR / 'Features.csv',
        CSV_DIR / 'Concepts_features.csv',
    ]

    table_props = {
        'rdf:ID': 'grammaticon',
        'dc:title': 'Grammaticon',
        'dc:source': 'sources.bib',
        'dcat:accessURL': 'https://github.com/clld/grammaticon-data',
        'prov:wasGeneratedBy': [
            {'dc:title': 'python',
             'dc:description': platform.python_version()},
             # TODO: do a pip freeze on the venv
             # {'dc:title': 'python-packages',
             #  'dc:relation': 'requirements.txt'},
        ],
    }

    if HERE.joinpath('.git').is_dir():
        git_remote = None
        with open(HERE / '.git' / 'config', encoding='utf-8') as f:
            for line in f:
                if re.fullmatch(r'\s*\[\s*remote\s+"origin"\s*\]\s*', line):
                    break
            for line in f:
                if (m := re.fullmatch(r'\s*url\s*=\s*(\S+)\s*', line)):
                    git_remote = m.group(1)
                    break
                elif re.match(r'\s*\[', line):
                    break
        if git_remote:
            git_exe = shutil.which('git')
            assert git_exe is not None
            procresult = subprocess.run(
                [git_exe, '-C', str(HERE), 'describe', '--always', '--tags'],
                stdout=subprocess.PIPE, check=True, encoding='utf-8')
            git_description = procresult.stdout.strip()
            table_props["prov:wasDerivedFrom"] = [
                {
                    'dc:title': 'Repository',
                    'rdf:type': 'prov:Entity',
                    'rdf:about': git_remote,
                    'dc:created': git_description,
                },
            ]

    table_data = {}
    table_meta_data = TableGroup(
        at_props={
            "context": [
                "http://www.w3.org/ns/csvw",
                {"@language": "en"},
            ],
        },
        common_props=table_props,
    )

    # load data

    for raw_path in raw_tables:
        table_spec = RAW_TO_CSWV_MAP[raw_path.name]
        table_name = table_spec['name']
        columns = table_spec['columns']
        with open(raw_path, encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            missing = [col for col in columns if col not in header]
            assert not missing, f'{raw_path}: missing fields: {missing}'
            unknown_fields = [col for col in header if col not in columns]
            assert not unknown_fields, f'{raw_path}: unknown fields: {unknown_fields}'
            mapped_colnames = [columns[colname]['name'] for colname in header]
            table_data[table_name] = [
                {k: v for k, v in zip(mapped_colnames, row) if v}
                for row in reader]

        table = Table(
            url=table_name,
            common_props=table_spec.get('properties') or {})
        table.tableSchema.columns = list(
            map(Column.fromvalue, columns.values()))
        for col, target_table in table_spec.get('foreign-keys', {}).items():
            table.add_foreign_key(col, target_table, 'ID')
        table_meta_data.tables.append(table)

    collection_ids_by_name = {row['Name']: row['ID'] for row in table_data['collections.csv']}
    with open(RAW_DIR / 'dois.csv') as f:
        zenodo_ids = {
            collection_ids_by_name[row['Name']]: get_zenodo_no(row['DOI'])
            for row in read_csv(f)}

    collection_archives = {
        collection_id: get_zip_path(zenodo_no)
        for collection_id, zenodo_no in zenodo_ids.items()}
    if (missing_archives := [p for p in collection_archives.values() if not p.exists()]):
        print('collections missing in download folder:', file=sys.stderr)
        print('\n'.join(f' * {p}' for p in missing_archives), file=sys.stderr)
        print('run `python3', sys.argv[0], 'download-collections` to download them', file=sys.stderr)
        sys.exit(66)

    collection_parameters = {
        collection_id: get_collection_parameters_from_zip(path)
        for collection_id, path in collection_archives.items()}

    # deal with the concept hierarchy separately

    with open(CSV_DIR / 'Concepthierarchy.csv', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = list(next(reader))
        original_hierarchy = [
            {k: v for k, v in zip(header, row) if v}
            for row in reader]

    sources = parse_file(RAW_DIR / 'sources.bib')

    # split the references
    for concept in table_data['concepts.csv']:
        if (source := concept.get('Source')):
            concept['Source'] = re.split(r'\s*;\s*', source)

    # add the data from the cldf datasets
    for feature in table_data['features.csv']:
        collection_id = feature.get('Collection_ID')
        id_in_collection = feature.get('ID_in_Collection')
        if collection_id and id_in_collection:
            collparams = collection_parameters[collection_id]
            collparam = collparams.get(id_in_collection) or {}
            feature['Language_Count'] = collparam.get('Language_Count') or 0
            feature['Name'] = feature.get('Name') or collparam['Name']

    table = Table(url='concept-hierarchy.csv')
    table.tableSchema.columns = [
        Column.fromvalue({'name': 'Child_ID', 'datatype': 'string'}),
        Column.fromvalue({'name': 'Parent_ID', 'datatype': 'string'})]
    table.add_foreign_key('Child_ID', 'concepts.csv', 'ID')
    table.add_foreign_key('Parent_ID', 'concepts.csv', 'ID')
    table_meta_data.tables.append(table)

    # ensure valid data

    for row in table_data['concepts.csv']:
        if (refs := row.get('Source')):
            row['Source'] = [BIBKEY_FIXES.get(key) or key for key in refs]

    bibkeys = {
        re.fullmatch(r'([^[]+)(?:\[[^\]]*\])?', citation).group(1).lower()
        for row in table_data['concepts.csv']
        for citation in row.get('Source') or ()}
    missing_bibkeys = {
        bibkey
        for bibkey in bibkeys
        if bibkey not in sources.entries}
    if missing_bibkeys:
        msg = '\n'.join(
            f'bibkey not found in bibliography: {bibkey}'
            for bibkey in sorted(missing_bibkeys))
        print(msg, file=sys.stderr)

    sources = BibliographyData(
        entries=sources.entries.__class__(
            (k, b)
            for k, b in sources.entries.items()
            if k.lower() in bibkeys))

    collection_ids = {r['ID'] for r in table_data['collections.csv']}

    table_data['concepts.csv'] = only_valid_concepts(
        table_data['concepts.csv'])
    table_data['features.csv'] = only_valid_features(
        table_data['features.csv'], collection_ids)

    concept_ids = {row['ID'] for row in table_data['concepts.csv']}
    feature_ids = {row['ID'] for row in table_data['features.csv']}

    table_data['concepts-features.csv'] = only_valid_concept_features(
        table_data['concepts-features.csv'], concept_ids, feature_ids)

    table_data['concept-hierarchy.csv'] = simplified_concept_hierarchy(
        original_hierarchy, concept_ids)

    # write data

    DEST_DIR.mkdir(parents=True, exist_ok=True)
    # clear out csvw folder
    for p in DEST_DIR.iterdir():
        p.unlink()
    table_meta_data.write(DEST_DIR / 'csvw-metadata.json', **table_data)
    sources.to_file(str(DEST_DIR / 'sources.bib'), 'bibtex')


def main():
    args = sys.argv
    if len(args) < 2:
        print(USAGE.format(progname=args[0]), file=sys.stderr)
        sys.exit(64)
    elif args[1] == 'xlsx-to-csv':
        xlsx_to_csv()
    elif args[1] == 'download-collections':
        download_collections()
    elif args[1] == 'make-csvw':
        make_csvw()
    elif args[1] in {'-h', '--help'}:
        print(USAGE.format(progname=args[0]), file=sys.stderr)
        sys.exit(64)
    else:
        print('Invalid command:', args[1], file=sys.stderr)
        print(USAGE.format(progname=args[0]), file=sys.stderr)
        sys.exit(64)


if __name__ == '__main__':
    main()
