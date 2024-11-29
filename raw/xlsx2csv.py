#!/usr/bin/env python3

import csv
from itertools import chain, repeat
from pathlib import Path

from openpyxl import load_workbook


def normalise_cell(value):
    if value is None:
        return ''
    else:
        return str(value).strip()


def pad_list(ls, width):
    if len(ls) == width:
        return ls
    elif len(ls) < width:
        return list(chain(ls, repeat('', width - len(ls))))
    else:
        raise ValueError(f'too long: {ls}')


def xlsx2csv(excel_path, outdir):
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    worksheets = wb.worksheets
    assert len(worksheets) == 1, f'{excel_path}: not exactly 1 worksheet'
    sheet = worksheets[0]
    rows = [
        row_norm
        for row in sheet.iter_rows()
        if any(row_norm := [normalise_cell(cell.value) for cell in row])]

    dest = outdir.joinpath(excel_path.name).with_suffix('.csv')
    table_width = max(len(row) for row in rows)
    if not outdir.is_dir():
        outdir.mkdir()
    with open(dest, 'w', encoding='utf-8') as f:
        wtr = csv.writer(f)
        wtr.writerows(pad_list(row, table_width) for row in rows)


def main():
    here = Path(__file__).parent
    outdir = here / 'csv-export'
    for p in here.glob('*.xlsx'):
        xlsx2csv(p, outdir)


if __name__ == '__main__':
    main()
